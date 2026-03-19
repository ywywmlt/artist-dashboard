"""
Project Justin — Korea 3 Shows Investor Model
FORMULA-DRIVEN version: edit any yellow input cell → everything recalculates.
Output: data/Project_Justin_Korea_Investor_Model.xlsx
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def f(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

THIN = Side(style="thin", color="CBD5E1")
ALL  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOT  = Border(bottom=THIN)
TOP  = Border(top=THIN)

INPUT_FILL   = fill("FFF9C4")   # yellow  — user editable
CALC_FILL    = fill("EFF6FF")   # light blue — formula calculated
HEADER_FILL  = fill("1E293B")   # dark
SECTION_FILL = fill("EEF2FF")   # indigo tint
SCEN_A_FILL  = fill("ECFDF5")   # green tint
SCEN_B_FILL  = fill("EDE9FE")   # violet tint
TOTAL_FILL   = fill("F1F5F9")   # grey
WHITE_FILL   = fill("FFFFFF")
ALT_FILL     = fill("F9FAFB")
RED_FILL     = fill("FEF2F2")
PDF_FILL     = fill("FFFBEB")

def cw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def rh(ws, row, height):
    ws.row_dimensions[row].height = height

def title(ws, r, c, text, end_col=None, bg="1E293B", fg="FFFFFF", sz=13):
    cell = ws.cell(row=r, column=c, value=text)
    cell.fill = fill(bg); cell.font = Font(bold=True, color=fg, size=sz)
    cell.alignment = al("left"); cell.border = ALL
    if end_col:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=end_col)
    rh(ws, r, 26)

def section(ws, r, c, text, end_col=None):
    cell = ws.cell(row=r, column=c, value=text)
    cell.fill = SECTION_FILL; cell.font = Font(bold=True, color="3730A3", size=10)
    cell.alignment = al("left"); cell.border = BOT
    if end_col:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=end_col)
    rh(ws, r, 18)

def hdr(ws, r, c, text, bg="334155", fg="FFFFFF"):
    cell = ws.cell(row=r, column=c, value=text)
    cell.fill = fill(bg); cell.font = Font(bold=True, color=fg, size=9)
    cell.alignment = al("center"); cell.border = ALL

def inp(ws, r, c, value, fmt="#,##0", label_col=None, label=""):
    """Yellow input cell"""
    if label_col:
        lc = ws.cell(row=r, column=label_col, value=label)
        lc.font = f(size=10); lc.alignment = al(); lc.border = BOT
    cell = ws.cell(row=r, column=c, value=value)
    cell.fill = INPUT_FILL; cell.font = Font(bold=True, color="1D4ED8", size=10)
    cell.alignment = al("right"); cell.number_format = fmt; cell.border = ALL
    return cell

def calc(ws, r, c, formula, fmt="#,##0", bold=False, color="000000", bg=None):
    """Blue calc cell"""
    cell = ws.cell(row=r, column=c, value=formula)
    cell.fill = bg if bg else CALC_FILL
    cell.font = Font(bold=bold, color=color, size=10)
    cell.alignment = al("right"); cell.number_format = fmt; cell.border = BOT
    return cell

def lbl(ws, r, c, text, bold=False, color="374151", bg=None, end_col=None, h="left"):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font = Font(bold=bold, color=color, size=10)
    cell.alignment = al(h); cell.border = BOT
    if bg: cell.fill = bg if isinstance(bg, PatternFill) else fill(bg)
    if end_col:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=end_col)
    return cell


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — INPUTS  (all yellow editable cells here)
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Inputs")
ws.sheet_view.showGridLines = False
for c, w in [(1,3),(2,36),(3,18),(4,18),(5,20)]:
    cw(ws, c, w)

r = 1
title(ws, r, 2, "INPUTS — PROJECT JUSTIN: KOREA 3 SHOWS", end_col=5,
      bg="1E293B", fg="FFFFFF", sz=13)
r += 1
title(ws, r, 2,
      "⚡ Edit YELLOW cells only — all other tabs recalculate automatically",
      end_col=5, bg="FEF3C7", fg="92400E", sz=10)
r += 2

# ── Section: Show Structure ───────────────────────────────────────────────────
section(ws, r, 2, "SHOW STRUCTURE", end_col=5); r += 1
hdr(ws, r, 2, "Parameter"); hdr(ws, r, 3, "Value"); hdr(ws, r, 4, "Unit"); hdr(ws, r, 5, "Notes")
r += 1

ws.cell(row=r, column=1, value="I5").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 3,        "#,##0", label_col=2, label="Number of Shows — Korea")
lbl(ws, r, 4, "shows"); lbl(ws, r, 5, "Korea stadium shows only", color="94A3B8"); r += 1

ws.cell(row=r, column=1, value="I6").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 40001,    "#,##0", label_col=2, label="Venue Capacity (per show)")
lbl(ws, r, 4, "seats"); lbl(ws, r, 5, "Korea stadium", color="94A3B8"); r += 1

ws.cell(row=r, column=1, value="I7").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, "2026-09-05", "@", label_col=2, label="Concert Start Date")
lbl(ws, r, 5, "5-Sep-26 confirmed", color="94A3B8"); r += 2

# ── Section: Seat Pricing & Counts ───────────────────────────────────────────
section(ws, r, 2, "SEAT PRICING & COUNTS — KOREA", end_col=5); r += 1
hdr(ws, r, 2, "Category"); hdr(ws, r, 3, "Price (USD)"); hdr(ws, r, 4, "Count (per show)"); hdr(ws, r, 5, "Notes")
r += 1

seat_data = [
    ("I10", "GA Standing (avg GA1–GA4)",        170,    19175, "General Admission standing"),
    ("I11", "Seating A — Lower Bowl",            250,    9588,  "Designated seating"),
    ("I12", "Seating B — Mid Deck",              300,    9588,  "Designated seating"),
    ("I13", "VIP Lounge",                        500,    1500,  "VIP floor lounge"),
    ("I14", "VVIP / Pre-Show Party Package",     1000,   150,   "Premium hospitality package"),
]
for ref, name, price, count, note in seat_data:
    ws.cell(row=r, column=1, value=ref).font = f(color="CBD5E1", size=8)
    lbl(ws, r, 2, name, bold=True)
    inp(ws, r, 3, price,  '"$"#,##0')
    inp(ws, r, 4, count,  "#,##0")
    lbl(ws, r, 5, note, color="94A3B8"); r += 1
r += 1

# ── Section: Additional Revenue ───────────────────────────────────────────────
section(ws, r, 2, "ADDITIONAL REVENUE INPUTS", end_col=5); r += 1
hdr(ws, r, 2, "Item"); hdr(ws, r, 3, "Value"); hdr(ws, r, 4, "Unit"); hdr(ws, r, 5, "Notes")
r += 1

ws.cell(row=r, column=1, value="I19").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 800000, '"$"#,##0', label_col=2, label="Expected Merch Sales (per show)")
lbl(ws, r, 4, "$/show"); lbl(ws, r, 5, "Total merch expected per show", color="94A3B8"); r += 1

ws.cell(row=r, column=1, value="I20").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 0.20, "0%", label_col=2, label="Promoter Merch Split")
lbl(ws, r, 4, "% of sales"); lbl(ws, r, 5, "Promoter receives 20%", color="94A3B8"); r += 1

ws.cell(row=r, column=1, value="I21").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 150000, '"$"#,##0', label_col=2, label="Sponsorship Revenue (per show)")
lbl(ws, r, 4, "$/show"); lbl(ws, r, 5, "Confirmed Korea shows", color="94A3B8"); r += 2

# ── Section: Artist MG ────────────────────────────────────────────────────────
section(ws, r, 2, "ARTIST FEE — MG PAYMENT SCHEDULE", end_col=5); r += 1
hdr(ws, r, 2, "Item"); hdr(ws, r, 3, "Value"); hdr(ws, r, 4, "Unit"); hdr(ws, r, 5, "Notes")
r += 1

ws.cell(row=r, column=1, value="I25").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 19500000, '"$"#,##0', label_col=2, label="Total MG — Justin")
lbl(ws, r, 4, "USD"); lbl(ws, r, 5, "Minimum Guarantee — contractual", color="94A3B8"); r += 1

ws.cell(row=r, column=1, value="I26").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 0.10, "0%", label_col=2, label="Payment 1 — % of MG (7 months before)")
lbl(ws, r, 4, "Feb-26"); r += 1

ws.cell(row=r, column=1, value="I27").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 0.40, "0%", label_col=2, label="Payment 2 — % of MG (6 months before)")
lbl(ws, r, 4, "Mar-26"); r += 1

ws.cell(row=r, column=1, value="I28").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 0.50, "0%", label_col=2, label="Payment 3 — % of MG (2 months before)")
lbl(ws, r, 4, "Jul-26"); r += 1

ws.cell(row=r, column=1, value="I29").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 0.00, "0%", label_col=2, label="Justin's Sponsorship Royalty")
lbl(ws, r, 4, "% of sponsorship"); r += 2

# ── Section: Scenario A ───────────────────────────────────────────────────────
section(ws, r, 2, "SCENARIO A — $12M EQUITY", end_col=5); r += 1
ws.cell(row=r, column=1, value="I32").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 250000, '"$"#,##0', label_col=2, label="Agency Fees")
lbl(ws, r, 4, "Jul-26"); lbl(ws, r, 5, "One-time, upfront", color="94A3B8"); r += 1

ws.cell(row=r, column=1, value="I33").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 250000, '"$"#,##0', label_col=2, label="BluFin Fees")
lbl(ws, r, 4, "Jul-26"); r += 1

ws.cell(row=r, column=1, value="I34").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 12000000, '"$"#,##0', label_col=2, label="Investor Equity")
lbl(ws, r, 4, "USD"); lbl(ws, r, 5, "Capital deployed Feb-26", color="94A3B8"); r += 1

ws.cell(row=r, column=1, value="I35").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 0.20, "0%", label_col=2, label="Hurdle Rate")
lbl(ws, r, 4, "%"); lbl(ws, r, 5, "20% minimum return threshold", color="94A3B8"); r += 1

ws.cell(row=r, column=1, value="I36").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 1224381, '"$"#,##0', label_col=2, label="Distribution after Hurdle — Investor")
lbl(ws, r, 5, "From full waterfall model", color="94A3B8"); r += 1

ws.cell(row=r, column=1, value="I37").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 1224381, '"$"#,##0', label_col=2, label="Distribution after Hurdle — BluFin")
lbl(ws, r, 5, "From full waterfall model", color="94A3B8"); r += 2

# ── Section: Scenario B ───────────────────────────────────────────────────────
section(ws, r, 2, "SCENARIO B — $25M EQUITY", end_col=5); r += 1
ws.cell(row=r, column=1, value="I40").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 500000, '"$"#,##0', label_col=2, label="Agency Fees")
lbl(ws, r, 4, "Jul-26"); r += 1

ws.cell(row=r, column=1, value="I41").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 500000, '"$"#,##0', label_col=2, label="BluFin Fees")
lbl(ws, r, 4, "Jul-26"); r += 1

ws.cell(row=r, column=1, value="I42").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 25000000, '"$"#,##0', label_col=2, label="Investor Equity")
lbl(ws, r, 4, "USD"); r += 1

ws.cell(row=r, column=1, value="I43").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 0.20, "0%", label_col=2, label="Hurdle Rate")
lbl(ws, r, 4, "%"); r += 1

ws.cell(row=r, column=1, value="I44").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 275131, '"$"#,##0', label_col=2, label="Distribution after Hurdle — Investor")
lbl(ws, r, 5, "From full waterfall model", color="94A3B8"); r += 1

ws.cell(row=r, column=1, value="I45").font = f(color="CBD5E1", size=8)
inp(ws, r, 3, 275131, '"$"#,##0', label_col=2, label="Distribution after Hurdle — BluFin")
lbl(ws, r, 5, "From full waterfall model", color="94A3B8"); r += 2

# ── Section: Net Operating Profit (monthly, from full model) ─────────────────
section(ws, r, 2, "NET OPERATING PROFIT — MONTHLY (from full model incl. variable costs)", end_col=5); r += 1
title(ws, r, 2, "These capture all operating costs (venue, production, staffing, tax etc.) "
      "Update when full cost model is available.", end_col=5, bg="FEF3C7", fg="92400E", sz=9)
r += 1
hdr(ws, r, 2, "Month"); hdr(ws, r, 3, "Scen A"); hdr(ws, r, 4, "Scen B"); hdr(ws, r, 5, "Notes")
r += 1

nop_data = [
    ("I49", "Feb-26 — Net Operating",    -542500,   -542500,  "Pre-show overhead"),
    ("I50", "May-26 — Net Operating",    17828134,  18669184, "Ticket sales month 1"),
    ("I51", "Jun-26 — Net Operating",    7640629,   8001079,  "Ticket sales month 2"),
    ("I52", "Jul-26 — Net Operating",    -1007500,  -1007500, "Fee payments + overhead"),
    ("I53", "Sep-26 — Net Operating",    930000,    930000,   "Concert month"),
]
for ref, month, va, vb, note in nop_data:
    ws.cell(row=r, column=1, value=ref).font = f(color="CBD5E1", size=8)
    lbl(ws, r, 2, month, bold=True)
    inp(ws, r, 3, va, '"$"#,##0;[Red]"($"#,##0")"')
    inp(ws, r, 4, vb, '"$"#,##0;[Red]"($"#,##0")"')
    lbl(ws, r, 5, note, color="94A3B8"); r += 1

r += 1
# Legend
title(ws, r, 2, "LEGEND:  🟡 Yellow = Input (you edit)   🔵 Blue = Formula (auto-calculates)   Do NOT edit blue cells",
      end_col=5, bg="F8FAFC", fg="475569", sz=9)


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — REVENUE  (formula-driven from Inputs)
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Revenue")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "C5"
for c, w in [(1,3),(2,32),(3,16),(4,16),(5,16),(6,16),(7,16)]:
    cw(ws, c, w)

r = 1
title(ws, r, 2, "REVENUE — KOREA 3 SHOWS", end_col=7)
r += 1
title(ws, r, 2, "All formulas pull from Inputs tab — change seat prices/counts there",
      end_col=7, bg="EFF6FF", fg="1D4ED8", sz=9)
r += 2

# ── Show overview ─────────────────────────────────────────────────────────────
section(ws, r, 2, "SHOW OVERVIEW", end_col=7); r += 1
hdr(ws, r, 2, "Metric"); hdr(ws, r, 3, "Show 1"); hdr(ws, r, 4, "Show 2"); hdr(ws, r, 5, "Show 3")
hdr(ws, r, 6, "Total (3 Shows)"); hdr(ws, r, 7, "Avg / Show")
r += 1

overview_rows = [
    ("Venue Capacity",      "=Inputs!C6",    "#,##0",          "059669"),
    ("Occupancy Rate",      "100%",          "0%",             "000000"),
    ("Expected Attendance", "=Inputs!C6",    "#,##0",          "000000"),
    ("Avg Ticket Price",    "",              '"$"#,##0.00',    "1D4ED8"),  # computed below
    ("Ticket Revenue",      "",              '"$"#,##0',       "059669"),  # computed below
]

# Capacity row
lbl(ws, r, 2, "Venue Capacity", bold=True)
for ci in [3,4,5]:
    calc(ws, r, ci, "=Inputs!C6", "#,##0", bold=True, color="059669")
calc(ws, r, 6, "=Inputs!C5*Inputs!C6", "#,##0", bold=True, color="059669")
calc(ws, r, 7, "=Inputs!C6", "#,##0", color="94A3B8")
r += 1

# Attendance row (same as capacity at 100%)
lbl(ws, r, 2, "Expected Attendance", bold=True)
for ci in [3,4,5]:
    calc(ws, r, ci, "=Inputs!C6", "#,##0")
calc(ws, r, 6, "=Inputs!C5*Inputs!C6", "#,##0", bold=True)
calc(ws, r, 7, "=Inputs!C6", "#,##0", color="94A3B8")
r += 2

# ── Seat Pricing & Revenue per category ──────────────────────────────────────
section(ws, r, 2, "SEAT PRICING × COUNTS → REVENUE", end_col=7); r += 1
hdr(ws, r, 2, "Category"); hdr(ws, r, 3, "Price (USD)"); hdr(ws, r, 4, "Seats / Show")
hdr(ws, r, 5, "Rev / Show"); hdr(ws, r, 6, "Total 3 Shows"); hdr(ws, r, 7, "% of Ticket Rev")
r += 1

# Price cells in Inputs: C10..C14, Count cells: D10..D14
price_refs = ["Inputs!C10","Inputs!C11","Inputs!C12","Inputs!C13","Inputs!C14"]
count_refs = ["Inputs!D10","Inputs!D11","Inputs!D12","Inputs!D13","Inputs!D14"]
cat_names  = ["GA Standing (avg GA1–GA4)","Seating A — Lower Bowl",
               "Seating B — Mid Deck","VIP Lounge","VVIP / Pre-Show Party Package"]

# Track the revenue per-show formula cells for summing
rev_per_show_cells = []
r_start_seat = r

for i, (cat, pr, cr) in enumerate(zip(cat_names, price_refs, count_refs)):
    bg = "F9FAFB" if i%2 else "FFFFFF"
    lbl(ws, r, 2, cat, bold=True, bg=bg)
    calc(ws, r, 3, f"={pr}", '"$"#,##0', bg=fill(bg))
    calc(ws, r, 4, f"={cr}", "#,##0", bg=fill(bg))
    rev_cell = f"E{r}"   # revenue per show
    calc(ws, r, 5, f"={pr}*{cr}", '"$"#,##0', color="1D4ED8", bg=fill(bg))
    rev_per_show_cells.append(f"E{r}")
    calc(ws, r, 6, f"=E{r}*Inputs!C5", '"$"#,##0', bold=True, color="059669", bg=fill(bg))
    r += 1

# Totals row
rev_sum = "+".join(rev_per_show_cells)
lbl(ws, r, 2, "TOTAL TICKET REVENUE (per show)", bold=True, bg="EEF2FF")
ws.cell(row=r, column=2).fill = SECTION_FILL
calc(ws, r, 3, f"=SUMPRODUCT({','.join(price_refs)},{','.join(count_refs)})/SUM({','.join(count_refs)})",
     '"$"#,##0.00', bold=True, color="1D4ED8", bg=SECTION_FILL)
ws.cell(row=r, column=3).number_format = '"$"#,##0.00'
ws.cell(row=r, column=3).value = f"=SUMPRODUCT({','.join(price_refs)},{','.join(count_refs)})/SUM({','.join(count_refs)})"
calc(ws, r, 4, f"=SUM({','.join(count_refs)})", "#,##0", bold=True, bg=SECTION_FILL)
ticket_per_show_row = r
calc(ws, r, 5, f"={'+'.join(rev_per_show_cells)}", '"$"#,##0', bold=True, color="059669", bg=SECTION_FILL)
calc(ws, r, 6, f"=E{r}*Inputs!C5", '"$"#,##0', bold=True, color="059669", bg=SECTION_FILL)
r += 2

# ── Additional Revenue ────────────────────────────────────────────────────────
section(ws, r, 2, "ADDITIONAL REVENUE", end_col=7); r += 1
hdr(ws, r, 2, "Stream"); hdr(ws, r, 3, "Per Show"); hdr(ws, r, 4, "Shows")
hdr(ws, r, 5, "Total"); hdr(ws, r, 6, "Notes")
r += 1

# Merch
lbl(ws, r, 2, "Expected Merch Sales (gross)", bold=True)
calc(ws, r, 3, "=Inputs!C19", '"$"#,##0')
calc(ws, r, 4, "=Inputs!C5",  "#,##0")
calc(ws, r, 5, "=C{0}*Inputs!C5".format(r), '"$"#,##0', color="94A3B8")
lbl(ws, r, 6, "Gross expected sales (not promoter revenue)", color="94A3B8"); r += 1

lbl(ws, r, 2, "  → Merch Revenue (promoter's share)", bold=True, color="059669")
calc(ws, r, 3, "=Inputs!C19*Inputs!C20", '"$"#,##0', color="059669")
calc(ws, r, 4, "=Inputs!C5",  "#,##0")
merch_total_cell = f"E{r}"
calc(ws, r, 5, "=C{0}*Inputs!C5".format(r), '"$"#,##0', bold=True, color="059669")
lbl(ws, r, 6, f"=Inputs!C20&\" of $\"&TEXT(Inputs!C19,\"#,##0\")&\" / show\"", color="94A3B8"); r += 1

lbl(ws, r, 2, "F&B Revenue", bold=True, color="94A3B8")
calc(ws, r, 3, 0, '"$"#,##0', color="94A3B8")
calc(ws, r, 4, "=Inputs!C5", "#,##0", color="94A3B8")
calc(ws, r, 5, 0, '"$"#,##0', color="94A3B8")
lbl(ws, r, 6, "Pending — buyer data required", color="DC2626"); r += 1

lbl(ws, r, 2, "Sponsorship Revenue", bold=True, color="059669")
calc(ws, r, 3, "=Inputs!C21", '"$"#,##0', color="059669")
calc(ws, r, 4, "=Inputs!C5", "#,##0")
sponsor_total_cell = f"E{r}"
calc(ws, r, 5, "=C{0}*Inputs!C5".format(r), '"$"#,##0', bold=True, color="059669")
lbl(ws, r, 6, "Confirmed Korea shows", color="94A3B8"); r += 1

# Grand Total
lbl(ws, r, 2, "TOTAL GROSS REVENUE", bold=True, bg="EEF2FF")
ws.cell(row=r, column=2).fill = SECTION_FILL
ticket_total_ref = f"F{ticket_per_show_row}"
calc(ws, r, 5, f"={ticket_total_ref}+{merch_total_cell}+{sponsor_total_cell}",
     '"$"#,##0', bold=True, color="059669", bg=SECTION_FILL)
lbl(ws, r, 6, "Tickets + Merch + Sponsorship", bold=True); r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 — ARTIST FEE
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Artist Fee")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B4"
for c, w in [(1,3),(2,36),(3,14),(4,14),(5,14),(6,22)]:
    cw(ws, c, w)

r = 1
title(ws, r, 2, "ARTIST FEE — MG PAYMENT SCHEDULE", end_col=6)
r += 1
title(ws, r, 2, "Change MG or % splits in Inputs tab → amounts recalculate automatically",
      end_col=6, bg="EFF6FF", fg="1D4ED8", sz=9)
r += 2

section(ws, r, 2, "MG OVERVIEW", end_col=6); r += 1
lbl(ws, r, 2, "Total MG — Justin", bold=True)
calc(ws, r, 3, "=Inputs!C25", '"$"#,##0', bold=True, color="DC2626")
lbl(ws, r, 6, "Source: Inputs!C25", color="94A3B8"); r += 1
lbl(ws, r, 2, "Concert Start Date", bold=True)
calc(ws, r, 3, "=Inputs!C7", "@", bold=True)
r += 1
lbl(ws, r, 2, "Sponsorship Royalty", bold=True)
calc(ws, r, 3, "=Inputs!C29", "0%", bold=True); r += 2

section(ws, r, 2, "PAYMENT SCHEDULE", end_col=6); r += 1
hdr(ws, r, 2, "Payment"); hdr(ws, r, 3, "Date"); hdr(ws, r, 4, "% of MG")
hdr(ws, r, 5, "Amount (USD)"); hdr(ws, r, 6, "Months Before Concert")
r += 1

pmt_rows = [
    ("1st Payment", "Feb-26", "=Inputs!C26", "=Inputs!C25*Inputs!C26", "7 months before"),
    ("2nd Payment", "Mar-26", "=Inputs!C27", "=Inputs!C25*Inputs!C27", "6 months before"),
    ("3rd Payment", "Jul-26", "=Inputs!C28", "=Inputs!C25*Inputs!C28", "2 months before"),
    ("4th Payment", "—",      "0%",          "—",                       "—"),
    ("5th Payment", "—",      "0%",          "—",                       "—"),
]
pmt_amount_cells = []
for i, (pname, pdate, ppct, pamount, ptiming) in enumerate(pmt_rows):
    bg = "F9FAFB" if i%2 else "FFFFFF"
    lbl(ws, r, 2, pname, bold=(i<3), bg=bg)
    lbl(ws, r, 3, pdate, bg=bg, h="center")
    c1 = ws.cell(row=r, column=4, value=ppct)
    c1.fill = CALC_FILL if ppct.startswith("=") else fill(bg)
    c1.font = Font(bold=(i<3), color="1D4ED8" if ppct.startswith("=") else "94A3B8", size=10)
    c1.alignment = al("right"); c1.number_format = "0%"; c1.border = BOT
    if pamount != "—":
        c2 = ws.cell(row=r, column=5, value=pamount)
        c2.fill = CALC_FILL; c2.font = Font(bold=True, color="DC2626", size=10)
        c2.alignment = al("right"); c2.number_format = '"$"#,##0'; c2.border = BOT
        pmt_amount_cells.append(f"E{r}")
    else:
        lbl(ws, r, 5, "—", h="center", bg=bg)
    lbl(ws, r, 6, ptiming, color="94A3B8", bg=bg); r += 1

# Total row
lbl(ws, r, 2, "TOTAL", bold=True, bg="EEF2FF"); ws.cell(row=r, column=2).fill = SECTION_FILL
lbl(ws, r, 3, "100%", h="center", bold=True, bg="EEF2FF"); ws.cell(row=r, column=3).fill = SECTION_FILL
calc(ws, r, 4, "=Inputs!C26+Inputs!C27+Inputs!C28", "0%", bold=True, bg=SECTION_FILL)
calc(ws, r, 5, f"={'+'.join(pmt_amount_cells)}", '"$"#,##0', bold=True, color="DC2626", bg=SECTION_FILL)
r += 2

# Validation note
title(ws, r, 2, "✓ Validation: Total payments should equal MG Total",
      end_col=6, bg="F0FDF4", fg="065F46", sz=9)
r += 1
lbl(ws, r, 2, "MG Total (Inputs)")
calc(ws, r, 3, "=Inputs!C25", '"$"#,##0', bold=True, color="DC2626"); r += 1
lbl(ws, r, 2, "Sum of Payments")
calc(ws, r, 3, f"={'+'.join(pmt_amount_cells)}", '"$"#,##0', bold=True, color="DC2626"); r += 1
lbl(ws, r, 2, "Difference (should be $0)")
calc(ws, r, 3, f"=Inputs!C25-({'+'.join(pmt_amount_cells)})", '"$"#,##0;[Red]"($"#,##0")"', bold=True)


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 4 — INVESTOR SCENARIOS
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Investor Scenarios")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B4"
for c, w in [(1,3),(2,32),(3,20),(4,20),(5,20),(6,24)]:
    cw(ws, c, w)

r = 1
title(ws, r, 2, "INVESTOR SCENARIOS — SIDE BY SIDE", end_col=6)
r += 1
title(ws, r, 2, "Change equity / fees in Inputs tab → all returns recalculate automatically",
      end_col=6, bg="EFF6FF", fg="1D4ED8", sz=9)
r += 2

# Column headers
hdr(ws, r, 2, "Item", bg="334155")
hdr(ws, r, 3, "SCENARIO A — $12M", bg="065F46")
hdr(ws, r, 4, "SCENARIO B — $25M", bg="4C1D95")
hdr(ws, r, 5, "PDF DECK (Reference)", bg="78350F")
hdr(ws, r, 6, "Notes", bg="334155")
r += 1

def scen_row_f(ws, row, label, f_a, f_b, f_pdf, note="", fmt="#,##0", bold=False):
    lbl(ws, row, 2, label, bold=bold)
    for ci, formula, bg in [(3, f_a, SCEN_A_FILL),(4, f_b, SCEN_B_FILL),(5, f_pdf, PDF_FILL)]:
        c = ws.cell(row=row, column=ci, value=formula)
        c.fill = bg
        if isinstance(formula, str) and formula.startswith("="):
            c.fill = bg
            clr = "065F46" if ci==3 else ("4C1D95" if ci==4 else "92400E")
        else:
            clr = "374151" if formula in ("—",None) else ("065F46" if ci==3 else "4C1D95")
        c.font = Font(bold=bold, color=clr, size=10)
        c.alignment = al("right"); c.number_format = fmt; c.border = BOT
    lbl(ws, row, 6, note, color="94A3B8")

scen_row_f(ws, r, "Equity Invested",
           "=Inputs!C34", "=Inputs!C42", 25000000,
           "Total capital deployed", '"$"#,##0'); r += 1

scen_row_f(ws, r, "Agency Fees",
           "=Inputs!C32", "=Inputs!C40", "—",
           "Paid Jul-26", '"$"#,##0'); r += 1

scen_row_f(ws, r, "BluFin Fees",
           "=Inputs!C33", "=Inputs!C41", "—",
           "Paid Jul-26", '"$"#,##0'); r += 1

scen_row_f(ws, r, "Total Fees",
           "=Inputs!C32+Inputs!C33", "=Inputs!C40+Inputs!C41", "—",
           "Agency + BluFin", '"$"#,##0'); r += 1

scen_row_f(ws, r, "Hurdle Rate",
           "=Inputs!C35", "=Inputs!C43", 0.20,
           "Min return threshold", "0%"); r += 1

scen_row_f(ws, r, "Return of Capital",
           "=Inputs!C34", "=Inputs!C42", "—",
           "Full capital returned", '"$"#,##0'); r += 1

scen_row_f(ws, r, "Dist. after Hurdle — Investor",
           "=Inputs!C36", "=Inputs!C44", "—",
           "From full waterfall model", '"$"#,##0'); r += 1

scen_row_f(ws, r, "Dist. after Hurdle — BluFin",
           "=Inputs!C37", "=Inputs!C45", "—",
           "From full waterfall model", '"$"#,##0'); r += 1

# Total Investor Return = Equity + both distributions
r_tir = r
scen_row_f(ws, r, "Total Investor Return",
           "=Inputs!C34+Inputs!C36+Inputs!C37",
           "=Inputs!C42+Inputs!C44+Inputs!C45",
           "—",
           "Equity returned + distributions", '"$"#,##0', bold=True); r += 1

# Investor Profit = Total Return - Equity
r_profit = r
scen_row_f(ws, r, "Investor Profit",
           f"=C{r_tir}-Inputs!C34",
           f"=D{r_tir}-Inputs!C42",
           "—",
           "Net gain on investment", '"$"#,##0', bold=True); r += 1

# MOIC = Total Return / Equity
r_moic = r
scen_row_f(ws, r, "MOIC",
           f"=C{r_tir}/Inputs!C34",
           f"=D{r_tir}/Inputs!C42",
           1.26,
           "Multiple on Invested Capital", '0.00"x"', bold=True); r += 1

# Absolute Return = Profit / Equity
r_ret = r
scen_row_f(ws, r, "Absolute Return",
           f"=C{r_profit}/Inputs!C34",
           f"=D{r_profit}/Inputs!C42",
           0.262,
           "", "0.0%", bold=True); r += 2

# PDF reference box
section(ws, r, 2, "PDF INVESTOR DECK — REFERENCE ASSUMPTIONS", end_col=6); r += 1
for label, val, fmt in [
    ("Total Expected Attendance", 240000,        "#,##0"),
    ("Average Ticket Price",      260,            '"$"#,##0'),
    ("Expected Ticket Revenue",   62512125,       '"$"#,##0'),
    ("Absolute Return",           0.262,          "0.0%"),
    ("MOIC",                      1.26,           '0.00"x"'),
]:
    lbl(ws, r, 2, label)
    c = ws.cell(row=r, column=3, value=val)
    c.fill = PDF_FILL; c.font = Font(bold=True, color="92400E", size=10)
    c.alignment = al("right"); c.number_format = fmt; c.border = BOT; r += 1
lbl(ws, r, 2,
    "Note: PDF deck uses 240K attendance @ $260 avg (6-show basis). "
    "Scen A/B above are 3-show Korea model.",
    color="94A3B8"); ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 5 — CASH FLOW
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Cash Flow")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "C5"

months = ["Feb-26","Mar-26","Apr-26","May-26","Jun-26","Jul-26","Aug-26","Sep-26"]
for c, w in [(1,3),(2,30)] + [(i+3, 14) for i in range(len(months)+1)]:
    cw(ws, c, w)

r = 1
title(ws, r, 2, "CASH FLOW — MONTHLY RUNNING BALANCE", end_col=len(months)+3)
r += 1
title(ws, r, 2, "Edit net operating profit rows in Inputs tab → cash flow updates automatically",
      end_col=len(months)+3, bg="EFF6FF", fg="1D4ED8", sz=9)
r += 2

hdr(ws, r, 2, "Item", bg="334155")
for mi, m in enumerate(months):
    hdr(ws, r, mi+3, m, bg="334155")
r += 1

# Helper: map month name to col index
def mc(month_name):
    return months.index(month_name) + 3

# ── Scenario A ────────────────────────────────────────────────────────────────
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=len(months)+2)
c = ws.cell(row=r, column=2, value="SCENARIO A — $12M EQUITY")
c.fill = fill("065F46"); c.font = Font(bold=True, color="FFFFFF", size=11)
c.alignment = al("center"); rh(ws, r, 20); r += 1

# Row refs for Scenario A (matching Inputs row numbers)
# Investor Capital in: Feb only
cf_a_rows = {}

def cf_row(ws, row, label, month_vals, fmt='"$"#,##0;[Red]"($"#,##0")"', bold=False, color="374151", bg_fill=None):
    """month_vals: dict of month_name → formula or value"""
    lbl(ws, row, 2, label, bold=bold, bg=bg_fill or "FFFFFF")
    for mi, m in enumerate(months):
        v = month_vals.get(m, None)
        c = ws.cell(row=row, column=mi+3, value=v if v is not None else "-")
        if v is not None and v != "-":
            c.fill = bg_fill or CALC_FILL
            is_neg = (isinstance(v, str) and "Inputs!C4" in v) or False
            c.font = Font(bold=bold, color=color, size=10)
            c.number_format = fmt
        else:
            c.fill = fill("F9FAFB"); c.font = Font(color="CBD5E1", size=10); c.alignment = al("center")
        c.alignment = al("right"); c.border = BOT
    return row

# Investor Capital In
r_cap_a = r
lbl(ws, r, 2, "Investor's Capital", bold=True)
for mi, m in enumerate(months):
    val = "=Inputs!C34" if m == "Feb-26" else "-"
    c = ws.cell(row=r, column=mi+3, value=val)
    c.fill = CALC_FILL if val != "-" else fill("F9FAFB")
    c.font = Font(bold=True, color="059669" if val!= "-" else "CBD5E1", size=10)
    c.alignment = al("right" if val!="-" else "center")
    c.number_format = '"$"#,##0'; c.border = BOT
r += 1

# Net Operating Profit — from Inputs
r_nop_a = r
lbl(ws, r, 2, "Net Operating Profit", bold=True)
nop_a_map = {"Feb-26":"=Inputs!C49","May-26":"=Inputs!C50","Jun-26":"=Inputs!C51",
             "Jul-26":"=Inputs!C52","Sep-26":"=Inputs!C53"}
for mi, m in enumerate(months):
    v = nop_a_map.get(m, None)
    c = ws.cell(row=r, column=mi+3, value=v if v else "-")
    if v:
        c.fill = INPUT_FILL; c.font = Font(bold=True, color="1D4ED8", size=10)
        c.number_format = '"$"#,##0;[Red]"($"#,##0")"'
    else:
        c.fill = fill("F9FAFB"); c.font = Font(color="CBD5E1", size=10); c.alignment = al("center")
    c.alignment = al("right"); c.border = BOT
r += 1

# Justin's Fees (outflows)
r_jf_a = r
lbl(ws, r, 2, "Justin's Fees", bold=True)
jf_a = {"Feb-26":"=-Inputs!C25*Inputs!C26",
        "Mar-26":"=-Inputs!C25*Inputs!C27",
        "Jul-26":"=-Inputs!C25*Inputs!C28"}
for mi, m in enumerate(months):
    v = jf_a.get(m, None)
    c = ws.cell(row=r, column=mi+3, value=v if v else "-")
    if v:
        c.fill = CALC_FILL; c.font = Font(bold=True, color="DC2626", size=10)
        c.number_format = '"$"#,##0;[Red]"($"#,##0")"'
    else:
        c.fill = fill("F9FAFB"); c.font = Font(color="CBD5E1", size=10); c.alignment = al("center")
    c.alignment = al("right"); c.border = BOT
r += 1

# Agency Fees
r_af_a = r
lbl(ws, r, 2, "Agency Fees")
for mi, m in enumerate(months):
    v = "=-Inputs!C32" if m == "Jul-26" else None
    c = ws.cell(row=r, column=mi+3, value=v if v else "-")
    if v:
        c.fill = CALC_FILL; c.font = Font(bold=True, color="DC2626", size=10)
        c.number_format = '"$"#,##0;[Red]"($"#,##0")"'
    else:
        c.fill = fill("F9FAFB"); c.font = Font(color="CBD5E1", size=10); c.alignment = al("center")
    c.alignment = al("right"); c.border = BOT
r += 1

# BluFin Fees
r_bf_a = r
lbl(ws, r, 2, "BluFin Fees")
for mi, m in enumerate(months):
    v = "=-Inputs!C33" if m == "Jul-26" else None
    c = ws.cell(row=r, column=mi+3, value=v if v else "-")
    if v:
        c.fill = CALC_FILL; c.font = Font(bold=True, color="DC2626", size=10)
        c.number_format = '"$"#,##0;[Red]"($"#,##0")"'
    else:
        c.fill = fill("F9FAFB"); c.font = Font(color="CBD5E1", size=10); c.alignment = al("center")
    c.alignment = al("right"); c.border = BOT
r += 1

# Return of Capital
r_roc_a = r
lbl(ws, r, 2, "Return of Capital")
for mi, m in enumerate(months):
    v = "=-Inputs!C34/2" if m in ("Jul-26","Sep-26") else None
    c = ws.cell(row=r, column=mi+3, value=v if v else "-")
    if v:
        c.fill = CALC_FILL; c.font = Font(bold=True, color="7C3AED", size=10)
        c.number_format = '"$"#,##0;[Red]"($"#,##0")"'
    else:
        c.fill = fill("F9FAFB"); c.font = Font(color="CBD5E1", size=10); c.alignment = al("center")
    c.alignment = al("right"); c.border = BOT
r += 1

# Total Cash Flow (running balance) — sum all above rows for each month
r_tcf_a = r
lbl(ws, r, 2, "Total Cash Flow (Running Balance)", bold=True, bg="EEF2FF")
ws.cell(row=r, column=2).fill = SECTION_FILL
for mi, m in enumerate(months):
    col_letter = get_column_letter(mi+3)
    formula = (f"=SUM({col_letter}{r_cap_a}:{col_letter}{r_roc_a})")
    c = ws.cell(row=r, column=mi+3, value=formula)
    c.fill = SECTION_FILL; c.font = Font(bold=True, color="1D4ED8", size=10)
    c.alignment = al("right"); c.number_format = '"$"#,##0;[Red]"($"#,##0")"'; c.border = BOT
r += 2

# Scenario A Summary
section(ws, r, 2, "SCENARIO A — INVESTOR SUMMARY", end_col=len(months)+2); r += 1
sum_a_rows = [
    ("Distribution after 20% absolute return", None, None),
    ("  Investor",  "=Inputs!C36", "059669"),
    ("  BluFin",    "=Inputs!C37", "059669"),
    ("Total Investor Return", "=Inputs!C34+Inputs!C36+Inputs!C37", "059669"),
    ("Investor Profit",       f"=Inputs!C34+Inputs!C36+Inputs!C37-Inputs!C34", "059669"),
    ("MOIC",                  "(Inputs!C34+Inputs!C36+Inputs!C37)/Inputs!C34", "7C3AED"),
    ("Absolute Return",       "(Inputs!C34+Inputs!C36+Inputs!C37-Inputs!C34)/Inputs!C34", "7C3AED"),
]
for label, formula, clr in sum_a_rows:
    is_header = formula is None
    bg = "EEF2FF" if is_header else (SCEN_A_FILL if label not in ("MOIC","Absolute Return") else "EDE9FE")
    lbl(ws, r, 2, label, bold=not is_header, bg=bg if not is_header else "EEF2FF")
    if formula:
        fmt = '0.00"x"' if label=="MOIC" else ("0.0%" if "Return" in label and label!="Total Investor Return" else '"$"#,##0')
        c = ws.cell(row=r, column=3, value=f"={formula}" if not formula.startswith("=") else formula)
        c.fill = fill(bg) if isinstance(bg, str) else bg
        c.font = Font(bold=True, color=clr, size=10)
        c.alignment = al("right"); c.number_format = fmt; c.border = BOT
    r += 1

r += 2
# ── Scenario B (abbreviated — same structure) ─────────────────────────────────
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=len(months)+2)
c = ws.cell(row=r, column=2, value="SCENARIO B — $25M EQUITY")
c.fill = fill("4C1D95"); c.font = Font(bold=True, color="FFFFFF", size=11)
c.alignment = al("center"); rh(ws, r, 20); r += 1

scen_b_lines = [
    ("Investor's Capital",  {"Feb-26": "=Inputs!C42"}, "059669", '"$"#,##0'),
    ("Net Operating Profit",{"Feb-26":"=Inputs!C49","May-26":"=Inputs!C50",
                              "Jun-26":"=Inputs!C51","Jul-26":"=Inputs!C52","Sep-26":"=Inputs!C53"},
                             "1D4ED8", '"$"#,##0;[Red]"($"#,##0")"'),
    ("Justin's Fees",       {"Feb-26":"=-Inputs!C25*Inputs!C26",
                              "Mar-26":"=-Inputs!C25*Inputs!C27",
                              "Jul-26":"=-Inputs!C25*Inputs!C28"}, "DC2626", '"$"#,##0;[Red]"($"#,##0")"'),
    ("Agency Fees",         {"Jul-26":"=-Inputs!C40"}, "DC2626", '"$"#,##0;[Red]"($"#,##0")"'),
    ("BluFin Fees",         {"Jul-26":"=-Inputs!C41"}, "DC2626", '"$"#,##0;[Red]"($"#,##0")"'),
    ("Return of Capital",   {"Jul-26":"=-Inputs!C42/2","Sep-26":"=-Inputs!C42/2"}, "7C3AED", '"$"#,##0;[Red]"($"#,##0")"'),
]
b_row_refs = []
for label, month_vals, clr, fmt in scen_b_lines:
    b_row_refs.append(r)
    lbl(ws, r, 2, label, bold=(label in ("Investor's Capital","Total Cash Flow")))
    for mi, m in enumerate(months):
        v = month_vals.get(m, None)
        c = ws.cell(row=r, column=mi+3, value=v if v else "-")
        if v:
            is_input = label == "Net Operating Profit"
            c.fill = INPUT_FILL if is_input else CALC_FILL
            c.font = Font(bold=True, color=clr, size=10)
            c.number_format = fmt
        else:
            c.fill = fill("F9FAFB"); c.font = Font(color="CBD5E1", size=10); c.alignment = al("center")
        c.alignment = al("right"); c.border = BOT
    r += 1

# Total Cash Flow B
lbl(ws, r, 2, "Total Cash Flow (Running Balance)", bold=True, bg="EEF2FF")
ws.cell(row=r, column=2).fill = SECTION_FILL
for mi, m in enumerate(months):
    col_letter = get_column_letter(mi+3)
    formula = f"=SUM({col_letter}{b_row_refs[0]}:{col_letter}{b_row_refs[-1]})"
    c = ws.cell(row=r, column=mi+3, value=formula)
    c.fill = SECTION_FILL; c.font = Font(bold=True, color="4C1D95", size=10)
    c.alignment = al("right"); c.number_format = '"$"#,##0;[Red]"($"#,##0")"'; c.border = BOT
r += 2

# Scenario B Summary
section(ws, r, 2, "SCENARIO B — INVESTOR SUMMARY", end_col=len(months)+2); r += 1
for label, formula, clr in [
    ("  Investor",    "=Inputs!C44", "059669"),
    ("  BluFin",      "=Inputs!C45", "059669"),
    ("Total Return",  "=Inputs!C42+Inputs!C44+Inputs!C45", "059669"),
    ("Profit",        "=Inputs!C44+Inputs!C45", "059669"),
    ("MOIC",          "=(Inputs!C42+Inputs!C44+Inputs!C45)/Inputs!C42", "7C3AED"),
    ("Abs. Return",   "=(Inputs!C44+Inputs!C45)/Inputs!C42", "7C3AED"),
]:
    fmt = '0.00"x"' if label=="MOIC" else ("0.0%" if "Return" in label and label!="Total Return" else '"$"#,##0')
    lbl(ws, r, 2, label, bold=True, bg=SCEN_B_FILL)
    c = ws.cell(row=r, column=3, value=formula)
    c.fill = SCEN_B_FILL; c.font = Font(bold=True, color=clr, size=10)
    c.alignment = al("right"); c.number_format = fmt; c.border = BOT; r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 6 — COST ASSUMPTIONS
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Cost Assumptions")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B4"
for c, w in [(1,3),(2,38),(3,18),(4,18),(5,18),(6,22)]:
    cw(ws, c, w)

r = 1
title(ws, r, 2, "COST ASSUMPTIONS", end_col=6)
r += 1
title(ws, r, 2, "Known costs link to Inputs. Variable costs need Run 3 data to complete.",
      end_col=6, bg="EFF6FF", fg="1D4ED8", sz=9)
r += 2

section(ws, r, 2, "KNOWN / FIXED COSTS", end_col=6); r += 1
hdr(ws, r, 2, "Line Item"); hdr(ws, r, 3, "Total (3 Shows)"); hdr(ws, r, 4, "Scen A"); hdr(ws, r, 5, "Scen B"); hdr(ws, r, 6, "Notes")
r += 1

known = [
    ("Artist MG — Justin",  "=Inputs!C25",             "=Inputs!C25",             "=Inputs!C25",             "Contractual — Inputs!C25"),
    ("Agency Fees",          "=Inputs!C32",             "=Inputs!C32",             "=Inputs!C40",             "Jul-26 · Scen A vs B differs"),
    ("BluFin Fees",          "=Inputs!C33",             "=Inputs!C33",             "=Inputs!C41",             "Jul-26 · Scen A vs B differs"),
]
known_rows = []
for i, (label, ftot, fa, fb, note) in enumerate(known):
    bg = "F9FAFB" if i%2 else "FFFFFF"
    lbl(ws, r, 2, label, bold=True, bg=bg)
    for ci, formula in [(3,ftot),(4,fa),(5,fb)]:
        c = ws.cell(row=r, column=ci, value=formula)
        c.fill = CALC_FILL; c.font = Font(bold=True, color="DC2626", size=10)
        c.alignment = al("right"); c.number_format = '"$"#,##0'; c.border = BOT
    lbl(ws, r, 6, note, color="94A3B8", bg=bg)
    known_rows.append(r); r += 1

# Total known costs
lbl(ws, r, 2, "TOTAL KNOWN COSTS", bold=True, bg="EEF2FF"); ws.cell(row=r, column=2).fill = SECTION_FILL
for ci, cols in [(3,["C"]),(4,["D"]),(5,["E"])]:
    refs = [f"{cols[0]}{kr}" for kr in known_rows]
    c = ws.cell(row=r, column=ci, value=f"=SUM({','.join(refs)})")
    c.fill = SECTION_FILL; c.font = Font(bold=True, color="DC2626", size=10)
    c.alignment = al("right"); c.number_format = '"$"#,##0'; c.border = TOP
r += 2

section(ws, r, 2, "VARIABLE / PENDING COSTS  — Need Run 3 data", end_col=6); r += 1
hdr(ws, r, 2, "Line Item"); hdr(ws, r, 3, "Input Cell"); hdr(ws, r, 4, "Status"); hdr(ws, r, 6, "Notes")
r += 1

variable_items = [
    ("Venue Fees",                        "% of Gross Revenue"),
    ("Production (Stage, Sound, AV)",     "Quote from production co."),
    ("Advertising & Promotion",           "Local + digital marketing budget"),
    ("Part Time Staffing",                "Event staff, security, ushers"),
    ("Licenses & Permits",                "Local government permits"),
    ("Public Liability Insurance",        "Event insurance certificate"),
    ("Cancellation / Postponement Ins.",  "Cover gross sales profit"),
    ("Catering & Backstage",              "Artist + crew catering"),
    ("Goods & Service Tax",               "% of Gross Revenue — Korea rate"),
    ("Music Copyright — COMPASS",         "% of Gross Revenue"),
    ("Miscellaneous / Contingency",       "Buffer — typically 3–5% of total"),
]
for i, (label, note) in enumerate(variable_items):
    lbl(ws, r, 2, label, bg=RED_FILL)
    lbl(ws, r, 3, "—", h="center", bg=RED_FILL)
    lbl(ws, r, 4, "❌  Missing — enter in Run 3", color="DC2626", bg=RED_FILL)
    for ci in [5]:
        ws.cell(row=r, column=ci, value="—").fill = fill("FEF2F2")
    lbl(ws, r, 6, note, color="94A3B8", bg=RED_FILL); r += 1

r += 1
title(ws, r, 2,
      "NOTE: Variable costs above are currently embedded in the 'Net Operating Profit' figures "
      "in the Cash Flow tab (Inputs C49–C53). When individual cost data is available, "
      "those can be broken out here and the NOP inputs updated accordingly.",
      end_col=6, bg="FEF3C7", fg="92400E", sz=9)


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 7 — SUMMARY  (formula-driven from all other tabs)
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Summary")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B3"
for c, w in [(1,3),(2,32),(3,20),(4,20),(5,22)]:
    cw(ws, c, w)

r = 1
title(ws, r, 2, "SUMMARY — PROJECT JUSTIN: KOREA 3 SHOWS", end_col=5, sz=14)
rh(ws, r, 32); r += 1
title(ws, r, 2, "INTERNAL — NOT FOR DISTRIBUTION  ·  All figures auto-update from Inputs tab",
      end_col=5, bg="7F1D1D", fg="FCA5A5", sz=9); r += 2

section(ws, r, 2, "REVENUE", end_col=5); r += 1
hdr(ws, r, 2, "Item"); hdr(ws, r, 3, "Amount"); hdr(ws, r, 4, "Per Show"); hdr(ws, r, 5, "Notes")
r += 1

# These reference the Revenue tab calculations
summary_rev = [
    ("Ticket Revenue (3 shows)",   "='Revenue'!F22",  "='Revenue'!E22",  "3 shows × 40,001 seats"),
    ("Merchandise Revenue",         "='Revenue'!E26",  "='Revenue'!C26",  "20% promoter split"),
    ("F&B Revenue",                 0,                  0,                 "Pending"),
    ("Sponsorship Revenue",         "='Revenue'!E28",  "='Revenue'!C28",  "$150K per show"),
    ("TOTAL GROSS REVENUE",         "='Revenue'!E30",  "='Revenue'!E30/Inputs!C5", ""),
]
for i, (label, ftot, fper, note) in enumerate(summary_rev):
    is_total = label.startswith("TOTAL")
    bg = "EEF2FF" if is_total else ("F9FAFB" if i%2 else "FFFFFF")
    lbl(ws, r, 2, label, bold=is_total, bg=bg)
    for ci, val in [(3,ftot),(4,fper)]:
        c = ws.cell(row=r, column=ci, value=val)
        c.fill = CALC_FILL if isinstance(val,str) and val.startswith("=") else fill(bg)
        c.font = Font(bold=is_total, color="059669", size=10)
        c.alignment = al("right"); c.number_format = '"$"#,##0'; c.border = BOT
    lbl(ws, r, 5, note, color="94A3B8", bg=bg); r += 1

r += 1
section(ws, r, 2, "COSTS", end_col=5); r += 1
hdr(ws, r, 2, "Item"); hdr(ws, r, 3, "Scen A"); hdr(ws, r, 4, "Scen B"); hdr(ws, r, 5, "Notes")
r += 1

summary_cost = [
    ("Artist MG",          "=Inputs!C25",                     "=Inputs!C25",                      "Justin — contractual"),
    ("Agency Fees",         "=Inputs!C32",                     "=Inputs!C40",                      "Upfront Jul-26"),
    ("BluFin Fees",         "=Inputs!C33",                     "=Inputs!C41",                      "Jul-26"),
    ("Total Known Costs",   "=Inputs!C25+Inputs!C32+Inputs!C33", "=Inputs!C25+Inputs!C40+Inputs!C41", "Excl. variable costs"),
    ("Variable Costs",      "TBD",                             "TBD",                              "Run 3 needed"),
]
for i, (label, fa, fb, note) in enumerate(summary_cost):
    is_total = "Total" in label
    bg = "EEF2FF" if is_total else ("F9FAFB" if i%2 else "FFFFFF")
    lbl(ws, r, 2, label, bold=is_total, bg=bg)
    for ci, val in [(3,fa),(4,fb)]:
        c = ws.cell(row=r, column=ci, value=val)
        c.fill = CALC_FILL if isinstance(val,str) and val.startswith("=") else fill(bg)
        c.font = Font(bold=is_total, color="DC2626", size=10)
        c.alignment = al("right")
        c.number_format = '"$"#,##0' if isinstance(val, str) and val.startswith("=") else "@"
        c.border = BOT
    lbl(ws, r, 5, note, color="94A3B8", bg=bg); r += 1

r += 1
section(ws, r, 2, "INVESTOR RETURNS", end_col=5); r += 1
hdr(ws, r, 2, "Item"); hdr(ws, r, 3, "Scenario A — $12M"); hdr(ws, r, 4, "Scenario B — $25M"); hdr(ws, r, 5, "Notes")
r += 1

inv_rows = [
    ("Equity Invested",       "=Inputs!C34",  "=Inputs!C42",  "#,##0",        "000000"),
    ("Total Investor Return",  "='Investor Scenarios'!C14", "='Investor Scenarios'!D14", "#,##0", "059669"),
    ("Investor Profit",        "='Investor Scenarios'!C15", "='Investor Scenarios'!D15", "#,##0", "059669"),
    ("MOIC",                   "='Investor Scenarios'!C16", "='Investor Scenarios'!D16", '0.00"x"', "7C3AED"),
    ("Absolute Return",        "='Investor Scenarios'!C17", "='Investor Scenarios'!D17", "0.0%",   "7C3AED"),
]
for i, (label, fa, fb, fmt, clr) in enumerate(inv_rows):
    is_bold = label in ("MOIC","Absolute Return","Total Investor Return")
    bg = SCEN_A_FILL if i%2==0 else "F9FAFB"
    lbl(ws, r, 2, label, bold=is_bold)
    for ci, val, bg_ in [(3,fa,SCEN_A_FILL),(4,fb,SCEN_B_FILL)]:
        c = ws.cell(row=r, column=ci, value=val)
        c.fill = CALC_FILL; c.font = Font(bold=is_bold, color=clr, size=10)
        c.alignment = al("right"); c.number_format = fmt; c.border = BOT
    r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# TAB — INVESTOR VIEW  (clean 3-scenario sell-through presentation)
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Investor View")
ws.sheet_view.showGridLines = False

# Column widths
for c, w in [(1, 3), (2, 36), (3, 24), (4, 24), (5, 24)]:
    cw(ws, c, w)

# Styles specific to this sheet
IV_BG       = fill("FFFFFF")
IV_HDR_BG   = fill("111827")   # near-black header
IV_BASE_BG  = fill("EFF6FF")   # light blue highlight for Base Case column
IV_BASE_HDR = fill("1E40AF")   # darker blue for Base Case header
IV_LABEL_FT = Font(size=11, color="374151")
IV_VAL_FT   = Font(size=12, color="111827", bold=True)
IV_PCT_FT   = Font(size=12, color="1D4ED8", bold=True)
IV_MOIC_FT  = Font(size=14, color="065F46", bold=True)
IV_NOTE_FT  = Font(size=9, color="6B7280", italic=True)
IV_THIN     = Border(bottom=Side(style="thin", color="E5E7EB"))

r = 1

# ── Title block ──────────────────────────────────────────────────────────────
ws.merge_cells("B1:E1")
c = ws.cell(row=1, column=2, value="INVESTOR VIEW — PROJECT JUSTIN")
c.font = Font(bold=True, color="111827", size=16)
c.alignment = al("left", "center")
rh(ws, 1, 36)
r += 1

ws.merge_cells("B2:E2")
c = ws.cell(row=2, column=2, value="Pre-Tax Scenario Analysis  ·  Scenario A ($12M Equity)")
c.font = Font(size=10, color="6B7280", italic=True)
c.alignment = al("left")
rh(ws, 2, 20)
r += 2   # r = 4

# ── Column headers ───────────────────────────────────────────────────────────
for ci, (label, bg_) in [
    (2, ("", IV_HDR_BG)),
    (3, ("Conservative", IV_HDR_BG)),
    (4, ("Base Case", IV_BASE_HDR)),
    (5, ("Upside", IV_HDR_BG)),
]:
    c = ws.cell(row=r, column=ci, value=label)
    c.fill = bg_
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = al("center", "center")
    c.border = ALL
rh(ws, r, 28)
r += 1   # r = 5

# ── Sell-Through % row ───────────────────────────────────────────────────────
lbl(ws, r, 2, "Sell-Through Rate", bold=True, color="6B7280")
for ci, pct, is_base in [(3, 0.85, False), (4, 0.95, True), (5, 1.00, False)]:
    c = ws.cell(row=r, column=ci, value=pct)
    c.fill = IV_BASE_BG if is_base else IV_BG
    c.font = Font(bold=True, color="1D4ED8" if is_base else "374151", size=11)
    c.alignment = al("center"); c.number_format = "0%"; c.border = IV_THIN
rh(ws, r, 22)
r += 1   # r = 6

# Thin separator
rh(ws, r, 8)
r += 1   # r = 7

# ── Helper to write a data row ───────────────────────────────────────────────
def iv_row(ws, row, label, formulas, fmt="#,##0", font_override=None, is_bold=False):
    """Write label + 3 formula cells (Conservative / Base / Upside)."""
    lc = ws.cell(row=row, column=2, value=label)
    lc.font = Font(bold=is_bold, size=11, color="111827" if is_bold else "374151")
    lc.alignment = al("left", "center"); lc.border = IV_THIN
    for ci, formula in [(3, formulas[0]), (4, formulas[1]), (5, formulas[2])]:
        is_base = (ci == 4)
        c = ws.cell(row=row, column=ci, value=formula)
        c.fill = IV_BASE_BG if is_base else IV_BG
        if font_override:
            c.font = font_override
        else:
            c.font = Font(bold=is_bold, size=12 if is_bold else 11,
                          color="111827" if not is_base else "1E40AF")
        c.alignment = al("right", "center"); c.number_format = fmt; c.border = IV_THIN
    rh(ws, row, 26)

# ── Financial rows ───────────────────────────────────────────────────────────
# Gross Revenue = (Ticket Rev at 100% × sell-through%) + Merch + Sponsorship
# Revenue!F16 = ticket total (3 shows, 100%), Revenue!E21 = merch, Revenue!E23 = sponsorship
iv_row(ws, r, "Gross Revenue", [
    "=('Revenue'!F16*C5)+'Revenue'!E21+'Revenue'!E23",
    "=('Revenue'!F16*D5)+'Revenue'!E21+'Revenue'!E23",
    "=('Revenue'!F16*E5)+'Revenue'!E21+'Revenue'!E23",
], fmt='"$"#,##0')
r += 1   # r = 8

# Total Costs (Scenario A: MG + Agency + BluFin)
# Inputs!C26=MG, Inputs!C33=Agency(A), Inputs!C34=BluFin(A)
iv_row(ws, r, "Total Costs", [
    "=Inputs!C26+Inputs!C33+Inputs!C34",
    "=Inputs!C26+Inputs!C33+Inputs!C34",
    "=Inputs!C26+Inputs!C33+Inputs!C34",
], fmt='"$"#,##0')
r += 1   # r = 9

# Net Profit = Gross Revenue - Total Costs
iv_row(ws, r, "Net Profit", [
    "=C7-C8", "=D7-D8", "=E7-E8",
], fmt='"$"#,##0', is_bold=True)
r += 1   # r = 10

# Separator
rh(ws, r, 10)
r += 1   # r = 11

# ROI (Pre-Tax) = Net Profit / Total Costs
iv_row(ws, r, "ROI (Pre-Tax)", [
    "=C9/C8", "=D9/D8", "=E9/E8",
], fmt="0.0%", font_override=IV_PCT_FT)
r += 1   # r = 12

# Return on Equity (Scenario A) = Net Profit / Investor Equity
# Inputs!C35 = Investor Equity ($12M)
iv_row(ws, r, "Return on Equity (Scenario A)", [
    "=C9/Inputs!C35", "=D9/Inputs!C35", "=E9/Inputs!C35",
], fmt="0.0%", font_override=IV_PCT_FT)
r += 1   # r = 13

# Investor MOIC (Scenario A) = (Equity + Net Profit) / Equity
iv_row(ws, r, "Investor MOIC (Scenario A)", [
    "=(Inputs!C35+C9)/Inputs!C35",
    "=(Inputs!C35+D9)/Inputs!C35",
    "=(Inputs!C35+E9)/Inputs!C35",
], fmt='0.00"x"', font_override=IV_MOIC_FT)
r += 1   # r = 14

# Separator
rh(ws, r, 10)
r += 1   # r = 15

# Implied Attendance = Capacity × Shows × Sell-Through%
# Inputs!C7=capacity, Inputs!C6=shows
iv_row(ws, r, "Implied Attendance", [
    "=Inputs!C7*Inputs!C6*C5",
    "=Inputs!C7*Inputs!C6*D5",
    "=Inputs!C7*Inputs!C6*E5",
], fmt="#,##0")
r += 2   # r = 17

# ── Footer note ──────────────────────────────────────────────────────────────
ws.merge_cells(f"B{r}:E{r}")
note_cell = ws.cell(row=r, column=2,
    value='All figures pre-tax, based on '
          '=TEXT(Inputs!C5,"0")&" shows"')
# Can't embed formula in a merged note cleanly — use static reference approach
note_cell.value = "All figures pre-tax. Show count and venue capacity per Inputs tab."
note_cell.font = IV_NOTE_FT
note_cell.alignment = al("left", "center")
rh(ws, r, 20)
r += 1

ws.merge_cells(f"B{r}:E{r}")
note2 = ws.cell(row=r, column=2,
    value="Costs reflect Scenario A structure (MG + Agency + BluFin fees). "
          "Variable costs embedded in net operating figures.")
note2.font = Font(size=8, color="9CA3AF", italic=True)
note2.alignment = al("left", "center")

# Print area
ws.print_area = f"B1:E{r}"


# ── Save ──────────────────────────────────────────────────────────────────────
out = "data/Project_Justin_Korea_Investor_Model.xlsx"
wb.save(out)

# Tab order: Investor View prominent
tab_order = ["Inputs","Summary","Investor View","Revenue","Artist Fee",
             "Cost Assumptions","Investor Scenarios","Cash Flow"]
tab_colors = ["FEF3C7","E0E7FF","DBEAFE","DCFCE7","FEE2E2","FEE2E2","EDE9FE","DBEAFE"]
for i, name in enumerate(tab_order):
    wb[name].sheet_properties.tabColor = tab_colors[i]
wb.save(out)

import os
size = os.path.getsize(out)
print(f"✅  Saved: {out}  ({size/1024:.0f} KB)")
print(f"📋  Tabs: {[s.title for s in wb.worksheets]}")
print()
print("🟡  Yellow cells = inputs (edit these)")
print("🔵  Blue cells   = formulas (auto-calculate)")
print("✅  Edit anything in the Inputs tab → Summary, Revenue, Cash Flow all update")
